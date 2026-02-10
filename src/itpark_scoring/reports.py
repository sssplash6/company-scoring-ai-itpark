from __future__ import annotations

import csv
from pathlib import Path
from typing import List

from fpdf import FPDF
from openpyxl import Workbook

from .models import CompanyResult


class ReportWriter:
    def __init__(self, output_dir: Path):
        self.output_dir = output_dir
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def write_csv(self, result: CompanyResult) -> Path:
        path = self.output_dir / f"{result.company_name}_scorecard.csv"
        with path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            writer.writerow(["Criterion", "Category", "Score", "Max", "Weight", "Rationale"])
            for criterion in result.scorecard.criteria:
                writer.writerow([
                    criterion.name,
                    criterion.category,
                    criterion.score,
                    criterion.max_score,
                    criterion.weight,
                    criterion.rationale,
                ])
        return path

    def write_excel(self, result: CompanyResult) -> Path:
        path = self.output_dir / f"{result.company_name}_scorecard.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Scorecard"
        ws.append(["Criterion", "Category", "Score", "Max", "Weight", "Rationale"])
        for criterion in result.scorecard.criteria:
            ws.append([
                criterion.name,
                criterion.category,
                criterion.score,
                criterion.max_score,
                criterion.weight,
                criterion.rationale,
            ])
        wb.save(path)
        return path

    def write_pdf(self, result: CompanyResult) -> Path:
        path = self.output_dir / f"{result.company_name}_scorecard.pdf"
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=12)
        pdf.add_page()
        pdf.set_font("Helvetica", size=14)
        pdf.cell(0, 10, f"Company Scorecard: {result.company_name}", ln=1)
        pdf.set_font("Helvetica", size=11)
        pdf.cell(0, 8, f"Overall score: {result.scorecard.overall_score}", ln=1)
        pdf.cell(0, 8, f"Coverage: {result.scorecard.coverage}", ln=1)
        pdf.cell(0, 8, f"Confidence: {result.scorecard.confidence}", ln=1)
        if result.scorecard.flags:
            pdf.cell(0, 8, f"Flags: {', '.join(result.scorecard.flags)}", ln=1)
        pdf.ln(4)
        pdf.set_font("Helvetica", size=10)
        for criterion in result.scorecard.criteria:
            line = (
                f"{criterion.category} | {criterion.name}: {criterion.score}/{criterion.max_score} "
                f"(w={criterion.weight})"
            )
            pdf.multi_cell(0, 6, line)
            pdf.set_text_color(80, 80, 80)
            pdf.multi_cell(0, 6, f"  {criterion.rationale}")
            pdf.set_text_color(0, 0, 0)
        pdf.output(str(path))
        return path
